"""
テンプレート Excel ファイル生成スクリプト
実行: python timetable/create_templates.py
"""

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from pathlib import Path

TEMPLATE_DIR = Path(__file__).parent / "templates"
TEMPLATE_DIR.mkdir(exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
NOTE_FONT   = Font(color="666666", italic=True, size=9)
THIN        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")


def _h(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = THIN
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _note(ws, row, col, val, color="FFF9C4"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = NOTE_FONT
    c.alignment = LEFT
    c.border = THIN


def create_flow_template():
    wb = openpyxl.Workbook()

    # ── シート: フロー ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "フロー"
    ws.sheet_view.showGridLines = False

    # タイトル
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "製造フロー定義シート"
    t.font = Font(bold=True, size=14)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # 説明行
    ws.merge_cells("A2:D2")
    d = ws["A2"]
    d.value = (
        "【記入方法】 操作番号は連番(1,2,3...)。前操作番号は複数の場合カンマ区切り(例: 1,2)。"
        "  機器・時間の設定はアプリ画面(③)で行ってください。"
    )
    d.font = NOTE_FONT
    d.fill = PatternFill("solid", fgColor="EBF5FB")
    d.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 40

    # ヘッダー（4列構成）
    headers = [
        ("操作番号", 10), ("操作名", 28), ("前操作番号", 14), ("操作タイプ", 18),
    ]
    for col, (h, w) in enumerate(headers, 1):
        _h(ws, 3, col, h, w)
    ws.row_dimensions[3].height = 20

    # 操作タイプ選択肢（ドロップダウン）
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_type = DataValidation(
        type="list",
        formula1='"CHARGE,HEAT,COOL,REACTION,CONCENTRATE,FILTER,TRANSFER,WASH,SEPARATION,CRYSTALLIZATION,OTHER"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="入力エラー",
        error="リストから選択してください",
    )
    dv_type.sqref = "D4:D100"
    ws.add_data_validation(dv_type)

    # サンプルデータ（4列: 操作番号, 操作名, 前操作番号, 操作タイプ）
    sample_rows = [
        (1, "原料仕込み",    "",  "CHARGE"),
        (2, "加熱昇温",      "1", "HEAT"),
        (3, "反応",          "2", "REACTION"),
        (4, "冷却",          "3", "COOL"),
        (5, "晶析",          "4", "CRYSTALLIZATION"),
        (6, "ろ過",          "5", "FILTER"),
        (7, "洗浄",          "6", "WASH"),
        (8, "濃縮",          "7", "CONCENTRATE"),
        (9, "移液・仕上げ",  "8", "TRANSFER"),
    ]
    fill_colors = ["FDFEFE", "FEF9E7"]
    for r, row in enumerate(sample_rows, start=4):
        fill = PatternFill("solid", fgColor=fill_colors[r % 2])
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.font = Font(size=10)
            cell.alignment = CENTER if c not in (2,) else LEFT
            cell.border = THIN
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A4"

    out_path = TEMPLATE_DIR / "flow_template.xlsx"
    wb.save(out_path)
    print(f"テンプレート生成完了: {out_path}")
    return out_path


if __name__ == "__main__":
    create_flow_template()
